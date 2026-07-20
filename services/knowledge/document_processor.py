"""Document Processing Service for Knowledge Space.

Author: lycosa9527
Made by: MindSpring Team

Extracts text from various file types including PDF, DOCX, images with OCR.

Copyright 2024-2025 北京思源智教科技有限公司 (Beijing Siyuan Zhijiao Technology Co., Ltd.)
All Rights Reserved
Proprietary License
"""

import logging
import mimetypes
import re
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from services.knowledge import document_ocr
from services.knowledge.audio_transcription import AUDIO_EXTENSIONS, transcribe_audio_file
from services.knowledge.document_office_extract import (
    extract_csv_markdown,
    extract_docx_markdown,
    extract_pptx_markdown,
    extract_xlsx_markdown,
)
from services.knowledge.legacy_office_convert import convert_legacy_office, is_legacy_office_mime
from services.utils.error_types import FILE_IO_ERRORS

logger = logging.getLogger(__name__)


class _DocumentProcessorState:
    """Holds the global DocumentProcessor singleton."""

    instance: Optional["DocumentProcessor"] = None


_document_processor_state = _DocumentProcessorState()

# Try to import language detection library
_langdetect_fn: Any = None
_langdetect_available = False
try:
    from langdetect import detect as _langdetect_detect

    _langdetect_fn = _langdetect_detect
    _langdetect_available = True
except ImportError:
    logger.warning("[DocumentProcessor] langdetect not available, language detection disabled")

# Try to import optional dependencies
_pypdf_mod: Any = None
_pypdf_available = False
try:
    import pypdf as _pypdf_import

    _pypdf_mod = _pypdf_import
    _pypdf_available = True
except ImportError:
    pass

_pdfplumber_mod: Any = None
_pdfplumber_available = False
try:
    import pdfplumber as _pdfplumber_import

    _pdfplumber_mod = _pdfplumber_import
    _pdfplumber_available = True
except ImportError:
    pass

_docx_document_cls: Any = None
_docx_available = False
try:
    from docx import Document as _docx_document_import

    _docx_document_cls = _docx_document_import
    _docx_available = True
except ImportError:
    pass

_pptx_presentation_cls: Any = None
_pptx_available = False
try:
    from pptx import Presentation as _pptx_presentation_import

    _pptx_presentation_cls = _pptx_presentation_import
    _pptx_available = True
except ImportError:
    pass

_openpyxl_load_workbook: Any = None
_openpyxl_available = False
try:
    from openpyxl import load_workbook as _openpyxl_load_workbook_import

    _openpyxl_load_workbook = _openpyxl_load_workbook_import
    _openpyxl_available = True
except ImportError:
    pass


class DocumentProcessor:
    """
    Document processing service for text extraction.

    Supports: PDF, DOCX/DOC, TXT, MD, CSV, images (with OCR), PPTX/PPT, XLSX/XLS, audio

    Includes file content validation using magic bytes to ensure files match their claimed type.
    """

    # Magic bytes (file signatures) for content validation
    # Format: (magic_bytes, mime_type, description)
    FILE_SIGNATURES = [
        # PDF
        (b"%PDF", "application/pdf", "PDF"),
        # Legacy OLE Office (.doc / .ppt / .xls)
        (b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1", "application/msword", "OLE Compound"),
        # DOCX (ZIP-based format, starts with PK)
        (
            b"PK\x03\x04",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "DOCX",
        ),
        # PPTX (ZIP-based format)
        (
            b"PK\x03\x04",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "PPTX",
        ),
        # XLSX (ZIP-based format)
        (
            b"PK\x03\x04",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "XLSX",
        ),
        # Images
        (b"\xff\xd8\xff", "image/jpeg", "JPEG"),
        (b"\x89PNG\r\n\x1a\n", "image/png", "PNG"),
        (b"RIFF", "image/webp", "WEBP"),
        (b"GIF87a", "image/gif", "GIF87a"),
        (b"GIF89a", "image/gif", "GIF89a"),
        (b"BM", "image/bmp", "BMP"),
        (b"II*\x00", "image/tiff", "TIFF (little-endian)"),
        (b"MM\x00*", "image/tiff", "TIFF (big-endian)"),
        # Text files (no magic bytes, but we can check for UTF-8 BOM)
        (b"\xef\xbb\xbf", "text/plain", "UTF-8 BOM"),
    ]

    MIME_DOCX = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    MIME_PPTX = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    MIME_DOC = "application/msword"
    MIME_PPT = "application/vnd.ms-powerpoint"
    MIME_XLS = "application/vnd.ms-excel"

    # Maximum magic bytes length to read
    MAX_MAGIC_LEN = max(len(sig[0]) for sig in FILE_SIGNATURES)

    def __init__(self):
        """Initialize document processor."""
        self.supported_types = {
            "application/pdf": self._extract_pdf,
            self.MIME_DOCX: self._extract_docx,
            self.MIME_DOC: self._extract_legacy_office,
            "text/plain": self._extract_text,
            "text/markdown": self._extract_text,
            "text/csv": self._extract_csv,
            "application/csv": self._extract_csv,
            "image/jpeg": self._extract_image,
            "image/png": self._extract_image,
            "image/jpg": self._extract_image,
            "image/webp": self._extract_image,
            self.MIME_PPTX: self._extract_pptx,
            self.MIME_PPT: self._extract_legacy_office,
            self.MIME_XLSX: self._extract_xlsx,
            self.MIME_XLS: self._extract_legacy_office,
        }
        # Audio sources are transcribed to text via DashScope recording-file ASR.
        for audio_mime in set(AUDIO_EXTENSIONS.values()):
            self.supported_types[audio_mime] = self._extract_audio

    def validate_file_content(self, file_path: str, expected_mime_type: str) -> Tuple[bool, Optional[str]]:
        """
        Validate file content using magic bytes (file signatures).

        Args:
            file_path: Path to file
            expected_mime_type: Expected MIME type

        Returns:
            Tuple of (is_valid, detected_mime_type)
            - is_valid: True if file content matches expected type
            - detected_mime_type: Detected MIME type from magic bytes (None if not detected)
        """
        try:
            with open(file_path, "rb") as f:
                header = f.read(self.MAX_MAGIC_LEN)

            if not header:
                return False, None

            # Check magic bytes
            detected_type = None
            for magic_bytes, mime_type, _description in self.FILE_SIGNATURES:
                if header.startswith(magic_bytes):
                    detected_type = mime_type
                    break

            # WEBP is RIFF....WEBP — confirm the WEBP marker at offset 8.
            if header.startswith(b"RIFF") and len(header) >= 12 and header[8:12] == b"WEBP":
                detected_type = "image/webp"

            # Special handling for ZIP-based formats (DOCX, PPTX, XLSX)
            # They all start with PK\x03\x04, so we need to check the internal structure
            if header.startswith(b"PK\x03\x04"):
                # Check for Office Open XML structure
                # DOCX contains word/document.xml
                # PPTX contains ppt/presentation.xml
                # XLSX contains xl/workbook.xml
                try:
                    with zipfile.ZipFile(file_path, "r") as zip_file:
                        file_list = zip_file.namelist()

                        # Check for DOCX
                        if any("word/document.xml" in f for f in file_list):
                            detected_type = self.MIME_DOCX
                        # Check for PPTX
                        elif any("ppt/presentation.xml" in f for f in file_list):
                            detected_type = self.MIME_PPTX
                        # Check for XLSX
                        elif any("xl/workbook.xml" in f for f in file_list):
                            detected_type = self.MIME_XLSX
                        else:
                            # It's a ZIP but not a recognized Office format
                            detected_type = "application/zip"
                except FILE_IO_ERRORS as e:
                    logger.warning("[DocumentProcessor] Failed to inspect ZIP structure: %s", e)
                    # If ZIP inspection fails, assume it matches if expected type is ZIP-based
                    if expected_mime_type in [self.MIME_DOCX, self.MIME_PPTX, self.MIME_XLSX]:
                        detected_type = expected_mime_type

            # OLE compound documents share one signature; trust the claimed legacy MIME.
            if header.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1") and is_legacy_office_mime(expected_mime_type):
                detected_type = expected_mime_type

            # For text files, we can't reliably detect from magic bytes alone
            # But we can check if it's valid UTF-8
            if expected_mime_type in ["text/plain", "text/markdown", "text/csv", "application/csv"]:
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        f.read(1024)  # Try to read first 1KB
                    detected_type = expected_mime_type  # Assume valid if readable as UTF-8
                except UnicodeDecodeError:
                    # CSV/text may be GBK; allow when extension-backed extractors will retry.
                    if expected_mime_type in ["text/csv", "application/csv", "text/plain"]:
                        detected_type = expected_mime_type
                    else:
                        return False, None

            # Validate match
            if detected_type:
                # Allow some flexibility for image types
                if expected_mime_type in ["image/jpeg", "image/jpg"] and detected_type == "image/jpeg":
                    return True, detected_type
                if expected_mime_type == detected_type:
                    return True, detected_type
                logger.warning(
                    "[DocumentProcessor] File content mismatch: expected %s, detected %s",
                    expected_mime_type,
                    detected_type,
                )
                return False, detected_type

            # If no magic bytes match but file exists, allow it (might be text or other format)
            # But log a warning
            logger.warning(
                "[DocumentProcessor] Could not detect file type from magic bytes: %s",
                file_path,
            )
            return True, None  # Allow processing but warn

        except FILE_IO_ERRORS as e:
            logger.error("[DocumentProcessor] Failed to validate file content: %s", e)
            return False, None

    def extract_references(self, text: str, document_id: int) -> List[Dict[str, Any]]:
        """
        Extract references and citations from document text.

        Args:
            text: Document text
            document_id: Document ID

        Returns:
            List of reference dicts: [{"type": "citation", "text": "...", "position": 123}, ...]
        """
        references = []

        # Extract citation patterns (e.g., [1], [Smith 2020], (Smith et al., 2020))
        citation_patterns = [
            r"\[(\d+)\]",  # [1], [2], etc.
            r"\(([A-Z][a-z]+(?:\s+et\s+al\.)?,\s+\d{4})\)",  # (Smith, 2020) or (Smith et al., 2020)
            r"\[([A-Z][a-z]+(?:\s+et\s+al\.)?,\s+\d{4})\]",  # [Smith, 2020]
        ]

        for pattern in citation_patterns:
            for match in re.finditer(pattern, text):
                references.append(
                    {
                        "type": "citation",
                        "text": match.group(0),
                        "position": match.start(),
                        "document_id": document_id,
                    }
                )

        # Extract cross-references (e.g., "see Section 3", "refer to Chapter 2")
        cross_ref_patterns = [
            r"(?:see|refer to|see also)\s+(?:section|chapter|figure|table|appendix)\s+(\d+)",
            r"(?:see|refer to|see also)\s+(?:Section|Chapter|Figure|Table|Appendix)\s+(\d+)",
        ]

        for pattern in cross_ref_patterns:
            for match in re.finditer(pattern, text, re.IGNORECASE):
                references.append(
                    {
                        "type": "cross_reference",
                        "text": match.group(0),
                        "position": match.start(),
                        "document_id": document_id,
                    }
                )

        return references

    def extract_metadata(self, file_path: str, file_type: str) -> Dict[str, Any]:
        """
        Extract metadata from document.

        Args:
            file_path: Path to file
            file_type: MIME type

        Returns:
            Dict with metadata: title, author, creation_date, etc.
        """
        if file_type not in self.supported_types:
            return {}

        try:
            if file_type == "application/pdf":
                return self._extract_pdf_metadata(file_path)
            if file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                return self._extract_docx_metadata(file_path)
            return {}
        except FILE_IO_ERRORS as e:
            logger.warning(
                "[DocumentProcessor] Failed to extract metadata from %s: %s",
                file_path,
                e,
            )
            return {}

    def _extract_pdf_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract metadata from PDF."""
        metadata = {}
        if not _pypdf_available or _pypdf_mod is None:
            return metadata
        try:
            with open(file_path, "rb") as file:
                pdf_reader = _pypdf_mod.PdfReader(file)
                if pdf_reader.metadata:
                    pdf_meta = pdf_reader.metadata
                    if pdf_meta.get("/Title"):
                        metadata["title"] = pdf_meta["/Title"]
                    if pdf_meta.get("/Author"):
                        metadata["author"] = pdf_meta["/Author"]
                    if pdf_meta.get("/CreationDate"):
                        metadata["creation_date"] = str(pdf_meta["/CreationDate"])
                    if pdf_meta.get("/Subject"):
                        metadata["subject"] = pdf_meta["/Subject"]
        except FILE_IO_ERRORS as e:
            logger.debug("[DocumentProcessor] PDF metadata extraction failed: %s", e)
        return metadata

    def _extract_docx_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract metadata from DOCX."""
        metadata = {}
        if not _docx_available or _docx_document_cls is None:
            return metadata
        try:
            doc = _docx_document_cls(file_path)
            core_props = doc.core_properties
            if core_props.title:
                metadata["title"] = core_props.title
            if core_props.author:
                metadata["author"] = core_props.author
            if core_props.created:
                metadata["creation_date"] = core_props.created.isoformat()
            if core_props.subject:
                metadata["subject"] = core_props.subject
            if core_props.keywords:
                metadata["keywords"] = core_props.keywords
        except FILE_IO_ERRORS as e:
            logger.debug("[DocumentProcessor] DOCX metadata extraction failed: %s", e)
        return metadata

    def detect_language(self, text: str) -> Optional[str]:
        """
        Detect language of text.

        Args:
            text: Text to analyze

        Returns:
            Language code (e.g., 'zh', 'en', 'ja') or None if detection fails
        """
        # Ensure text is a string
        if isinstance(text, list):
            logger.warning("[DocumentProcessor] detect_language received list, converting to string")
            text = "\n".join(str(item) for item in text)
        if not isinstance(text, str):
            text = str(text) if text else ""

        if not _langdetect_available or _langdetect_fn is None or not text or len(text.strip()) < 10:
            return None

        try:
            # Use first 1000 characters for faster detection
            sample = text[:1000] if len(text) > 1000 else text
            lang_code = _langdetect_fn(sample)
            logger.debug("[DocumentProcessor] Detected language: %s", lang_code)
            return lang_code
        except FILE_IO_ERRORS as exc:
            logger.debug("[DocumentProcessor] Language detection failed: %s", exc)
            return None

    def extract_text(self, file_path: str, file_type: str) -> str:
        """
        Extract text from file.

        Args:
            file_path: Path to file
            file_type: MIME type

        Returns:
            Extracted text content
        """
        if file_type not in self.supported_types:
            raise ValueError(f"Unsupported file type: {file_type}")

        # Validate file content matches claimed type
        is_valid, detected_type = self.validate_file_content(file_path, file_type)
        if not is_valid:
            if detected_type:
                raise ValueError(
                    f"File content does not match claimed type. Claimed: {file_type}, Detected: {detected_type}"
                )
            raise ValueError(f"File content validation failed. File may be corrupted or not match type: {file_type}")

        extractor = self.supported_types[file_type]
        try:
            text = extractor(file_path)
            # Handle case where extractor might return a list (shouldn't happen, but defensive)
            if isinstance(text, list):
                logger.warning(
                    "[DocumentProcessor] Extractor returned list instead of string for %s, joining",
                    file_path,
                )
                text = "\n".join(str(item) for item in text)
            # Ensure text is a string
            if not isinstance(text, str):
                text = str(text) if text else ""
            if not text or not text.strip():
                raise ValueError(f"No text extracted from {file_path}")
            return text.strip()
        except FILE_IO_ERRORS as e:
            logger.error("[DocumentProcessor] Failed to extract text from %s: %s", file_path, e)
            raise

    # Below this many extracted characters per page, treat the PDF as scanned
    # (image-only) and fall back to rasterized-page OCR.
    MIN_PDF_CHARS_PER_PAGE = 8
    # Bound OCR cost/latency for very large scanned PDFs.
    MAX_PDF_OCR_PAGES = 50
    # Rasterization DPI for scanned-page OCR (higher = sharper but slower).
    PDF_OCR_DPI = 200

    def _extract_pdf(self, file_path: str) -> str:
        """Extract text from PDF, falling back to OCR for scanned (image-only) PDFs."""
        text, page_count = self._extract_pdf_text_layer(file_path)
        if not self._pdf_text_is_sparse(text, page_count):
            return text

        logger.info(
            "[DocumentProcessor] PDF text layer sparse (%d chars, %d pages); attempting OCR: %s",
            len(text.strip()),
            page_count,
            file_path,
        )
        ocr_text, _ = self._ocr_pdf_pages(file_path)
        return ocr_text if len(ocr_text.strip()) > len(text.strip()) else text

    def _pdf_text_is_sparse(self, text: str, page_count: int) -> bool:
        """True when a PDF yields too little text to be anything but scanned images."""
        threshold = max(self.MIN_PDF_CHARS_PER_PAGE * max(page_count, 1), self.MIN_PDF_CHARS_PER_PAGE)
        return len(text.strip()) < threshold

    def _extract_pdf_text_layer(self, file_path: str) -> Tuple[str, int]:
        """Extract the embedded text layer from a PDF; returns (text, page_count)."""
        if not _pypdf_available and (not _pdfplumber_available or _pdfplumber_mod is None):
            raise ImportError("pypdf or pdfplumber required for PDF extraction")

        if _pypdf_available and _pypdf_mod is not None:
            try:
                text_parts = []
                with open(file_path, "rb") as file:
                    pdf_reader = _pypdf_mod.PdfReader(file)
                    page_count = len(pdf_reader.pages)
                    for page in pdf_reader.pages:
                        text = page.extract_text()
                        if text:
                            text_parts.append(text)
                return "\n\n".join(text_parts), page_count
            except FILE_IO_ERRORS as e:
                logger.warning("[DocumentProcessor] pypdf failed, trying pdfplumber: %s", e)
                if _pdfplumber_available and _pdfplumber_mod is not None:
                    return self._pdfplumber_text_layer(file_path)
                raise

        if _pdfplumber_available and _pdfplumber_mod is not None:
            return self._pdfplumber_text_layer(file_path)

        raise ImportError("pypdf or pdfplumber required for PDF extraction")

    def _pdfplumber_text_layer(self, file_path: str) -> Tuple[str, int]:
        """Extract text + page count via pdfplumber."""
        if not _pdfplumber_available or _pdfplumber_mod is None:
            raise ImportError("pdfplumber required for PDF extraction")
        with _pdfplumber_mod.open(file_path) as pdf:
            text_parts = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return "\n\n".join(text_parts), len(pdf.pages)

    def _ocr_pdf_pages(self, file_path: str) -> Tuple[str, List[Dict[str, Any]]]:
        """Rasterize and OCR a scanned PDF via the shared OCR module."""
        return document_ocr.ocr_pdf_pages(file_path, self.MAX_PDF_OCR_PAGES, self.PDF_OCR_DPI)

    def extract_text_with_pages(self, file_path: str, file_type: str) -> tuple[str, List[Dict[str, Any]]]:
        """
        Extract text with page information for PDFs.

        Args:
            file_path: Path to file
            file_type: MIME type

        Returns:
            Tuple of (text, page_info) where page_info is list of page boundaries
        """
        if file_type == "application/pdf":
            if not _pdfplumber_available or _pdfplumber_mod is None:
                # Fallback: extract text without page info
                text = self._extract_pdf(file_path)
                return text, []
            try:
                with _pdfplumber_mod.open(file_path) as pdf:
                    text_parts = []
                    page_info = []
                    current_pos = 0
                    page_total = len(pdf.pages)
                    for page_num, page in enumerate(pdf.pages, 1):
                        text = page.extract_text()
                        if text:
                            start_pos = current_pos
                            text_parts.append(text)
                            current_pos += len(text) + 2  # +2 for "\n\n"
                            page_info.append(
                                {
                                    "page": page_num,
                                    "start": start_pos,
                                    "end": current_pos,
                                }
                            )
                    joined = "\n\n".join(text_parts)
                # Scanned PDF: text layer is sparse → OCR rasterized pages.
                if self._pdf_text_is_sparse(joined, page_total):
                    logger.info(
                        "[DocumentProcessor] PDF text layer sparse (%d chars, %d pages); attempting OCR: %s",
                        len(joined.strip()),
                        page_total,
                        file_path,
                    )
                    ocr_text, ocr_pages = self._ocr_pdf_pages(file_path)
                    if len(ocr_text.strip()) > len(joined.strip()):
                        return ocr_text, ocr_pages
                return joined, page_info
            except ImportError:
                # Fallback: extract text without page info
                text = self._extract_pdf(file_path)
                return text, []
        else:
            # For non-PDF files, no page info
            text = self.extract_text(file_path, file_type)
            return text, []

    def _extract_docx(self, file_path: str) -> str:
        """Extract text and tables from DOCX as markdown."""
        if not _docx_available or _docx_document_cls is None:
            raise ImportError("python-docx required for DOCX extraction")
        return extract_docx_markdown(_docx_document_cls, file_path)

    def _extract_csv(self, file_path: str) -> str:
        """Extract a CSV file as a markdown table."""
        return extract_csv_markdown(file_path)

    def _extract_legacy_office(self, file_path: str) -> str:
        """Convert legacy Office binaries to OOXML, then extract."""
        mime_type = self.get_file_type(file_path)
        if not is_legacy_office_mime(mime_type):
            # Fallback: infer from extension when callers pass a temp without suffix.
            ext = Path(file_path).suffix.lower()
            mime_by_ext = {
                ".doc": self.MIME_DOC,
                ".ppt": self.MIME_PPT,
                ".xls": self.MIME_XLS,
            }
            mime_type = mime_by_ext.get(ext, mime_type)
        if not is_legacy_office_mime(mime_type):
            raise ValueError(f"Unsupported legacy Office type for {file_path}")

        out_dir = tempfile.mkdtemp(prefix="legacy_office_")
        try:
            converted_path, ooxml_mime = convert_legacy_office(file_path, mime_type, out_dir)
            extractor = self.supported_types.get(ooxml_mime)
            if extractor is None or extractor is self._extract_legacy_office:
                raise ValueError(f"No OOXML extractor for converted type {ooxml_mime}")
            return extractor(converted_path)
        finally:
            shutil.rmtree(out_dir, ignore_errors=True)

    def _extract_text(self, file_path: str) -> str:
        """Extract text from plain text files."""
        # Try multiple encodings
        encodings = ["utf-8", "gbk", "gb2312", "latin-1"]
        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding) as file:
                    return file.read()
            except UnicodeDecodeError:
                continue
            except FILE_IO_ERRORS as e:
                logger.warning("[DocumentProcessor] Failed to read with %s: %s", encoding, e)
                continue

        raise ValueError(f"Failed to decode text file with any encoding: {file_path}")

    def _extract_image(self, file_path: str) -> str:
        """Extract text from image using DashScope OCR (Tesseract fallback)."""
        return document_ocr.ocr_image_file(file_path, self.get_file_type(file_path))

    def _extract_audio(self, file_path: str) -> str:
        """Transcribe an audio source to text via DashScope recording-file ASR."""
        return transcribe_audio_file(file_path)

    def _extract_pptx(self, file_path: str) -> str:
        """Extract slide text and speaker notes from PPTX as markdown."""
        if not _pptx_available or _pptx_presentation_cls is None:
            raise ImportError("python-pptx required for PPTX extraction")
        return extract_pptx_markdown(_pptx_presentation_cls, file_path)

    def _extract_xlsx(self, file_path: str) -> str:
        """Extract workbook sheets from XLSX as markdown tables."""
        if not _openpyxl_available or _openpyxl_load_workbook is None:
            raise ImportError("openpyxl required for XLSX extraction")
        return extract_xlsx_markdown(_openpyxl_load_workbook, file_path)

    def get_file_type(self, file_path: str) -> str:
        """
        Get MIME type of file.

        Args:
            file_path: Path to file

        Returns:
            MIME type string
        """
        ext = Path(file_path).suffix.lower()
        # Canonicalize audio extensions first: mimetypes is inconsistent across
        # platforms (e.g. .wav → audio/x-wav, .m4a/.opus → None) and must match
        # the supported_types keys derived from AUDIO_EXTENSIONS.
        if ext in AUDIO_EXTENSIONS:
            return AUDIO_EXTENSIONS[ext]

        ext_to_mime = {
            ".pdf": "application/pdf",
            ".docx": self.MIME_DOCX,
            ".doc": self.MIME_DOC,
            ".txt": "text/plain",
            ".md": "text/markdown",
            ".csv": "text/csv",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".webp": "image/webp",
            ".pptx": self.MIME_PPTX,
            ".ppt": self.MIME_PPT,
            ".xlsx": self.MIME_XLSX,
            ".xls": self.MIME_XLS,
        }
        if ext in ext_to_mime:
            return ext_to_mime[ext]

        mime_type, _ = mimetypes.guess_type(file_path)
        return mime_type or "application/octet-stream"

    def is_supported(self, file_type: str) -> bool:
        """
        Check if file type is supported.

        Args:
            file_type: MIME type

        Returns:
            True if supported
        """
        return file_type in self.supported_types


def get_document_processor() -> DocumentProcessor:
    """Get global document processor instance."""
    if _document_processor_state.instance is None:
        _document_processor_state.instance = DocumentProcessor()
    return _document_processor_state.instance
